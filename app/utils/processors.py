from collections import defaultdict

from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
from dateutil import parser as dt_parser

REQUIRED_SHEET_NAME = "LBS"

REQUIRED_COLUMNS = {
    "Msisdn / Номер ISDN мобильного абонента",
    "Период времени",
    "Широта",
    "Долгота",
}

COLUMN_MAPPING = {
    "Source operator/Оператор источника": "source_operator",
    "Msisdn / Номер ISDN мобильного абонента": "isdn_number",
    "Период времени": "time_period",
    "IMSI": "imsi_number",
    "Тип записи": "record_type",
    "Тип события": "action_type",
    "LAC/TAC-CellId": "lac_tac",
    "Тип базовой станции": "base_station_type",
    "Azimuth / Азимут": "azimuth",
    "Широта": "width",
    "Долгота": "height",
    "Radius / Радиус действия базовой станции": "radius",
    "Base Station Location Address / Адрес базовой станции": "base_station_location",
    "Регион": "region",
    "IMEI": "imei"
}

FIELDS_FORCE_STR = {
    "isdn_number", "imsi_number", "lac_tac", "radius",
    "base_station_location", "region", "imei", "width", "height"
}

FIELDS_FORCE_FLOAT = {"azimuth"}

FIELDS_FORCE_DATETIME = {"time_period"}


def parse_excel_data(contents: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(contents), read_only=True)
    ws = wb[REQUIRED_SHEET_NAME]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {i: COLUMN_MAPPING.get(col, None) for i, col in enumerate(headers)}

    parsed_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        skip = False

        for i, value in enumerate(row):
            original_header = headers[i]
            mapped_key = header_map.get(i)

            if not mapped_key:
                continue

            value_str = str(value).strip() if value is not None else ""

            if original_header in REQUIRED_COLUMNS and value_str in ("", "-"):
                skip = True
                break

            if value_str in ("", "-"):
                row_dict[mapped_key] = None
            else:
                if mapped_key in FIELDS_FORCE_STR:
                    row_dict[mapped_key] = str(value)
                elif mapped_key in FIELDS_FORCE_DATETIME:
                    try:
                        dt_obj = value if isinstance(value, datetime) else dt_parser.parse(value_str)\

                        if dt_obj.tzinfo is not None:
                            dt_obj = dt_obj.replace(tzinfo=None)

                        row_dict[mapped_key] = dt_obj
                    except Exception:
                        skip = True
                        break
                elif mapped_key in FIELDS_FORCE_FLOAT:
                    try:
                        row_dict[mapped_key] = float(value)
                    except (ValueError, TypeError):
                        row_dict[mapped_key] = None
                else:
                    row_dict[mapped_key] = value

        if not skip:
            parsed_rows.append(row_dict)

    return parsed_rows




import pandas as pd

def parse_excel_data_fast(contents: bytes) -> list[dict]:
    df = pd.read_excel(BytesIO(contents), sheet_name="LBS", engine="openpyxl", dtype=str)

    # Заменяем '-' на None
    df.replace("-", None, inplace=True)

    skipped_rows = df[
        df[["Msisdn / Номер ISDN мобильного абонента", "Период времени", "Широта", "Долгота"]].isnull().any(axis=1)]

    # Удаляем строки без обязательных колонок
    df.dropna(subset=[
        "Msisdn / Номер ISDN мобильного абонента",
        "Период времени",
        "Широта",
        "Долгота"
    ], inplace=True)

    # Переименовываем
    df.rename(columns=COLUMN_MAPPING, inplace=True)

    # Типизация
    df["time_period"] = pd.to_datetime(df["time_period"], errors="coerce", utc=True)
    df["time_period"] = df["time_period"].dt.tz_localize(None)
    df["azimuth"] = pd.to_numeric(df.get("azimuth", None), errors="coerce")

    # Возврат в виде списка словарей
    return df.to_dict(orient="records")


def parse_excel_data_parallel(contents: bytes) -> dict[str, list[dict]]:
    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    ws = wb["LBS"]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    msisdn_col = header.index("Msisdn / Номер ISDN мобильного абонента")

    rows_by_number = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        number = row[msisdn_col]
        if not number:
            continue

        record = dict(zip(header, row))
        rows_by_number[number].append(record)

    return rows_by_number


def parse_excel_stream_grouped(contents: bytes) -> dict[str, list[dict]]:
    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    ws = wb["LBS"]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    number_idx = header.index("Msisdn / Номер ISDN мобильного абонента")

    groups = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[number_idx]:
            continue
        number = str(row[number_idx])
        record = dict(zip(header, row))
        groups[number].append(record)

    return groups
