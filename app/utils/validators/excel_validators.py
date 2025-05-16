from fastapi import UploadFile
from io import BytesIO

from openpyxl.reader.excel import load_workbook

from app.utils.errors import BadRequestException

REQUIRED_SHEET_NAME = "LBS"

REQUIRED_COLUMNS = {
    "Source operator/Оператор источника",
    "Msisdn / Номер ISDN мобильного абонента",
    "Период времени",
    "IMSI",
    "Тип записи",
    "Тип события",
    "LAC/TAC-CellId",
    "Тип базовой станции",
    "Azimuth / Азимут",
    "Широта",
    "Долгота",
    "Radius / Радиус действия базовой станции",
    "Base Station Location Address / Адрес базовой станции",
    "Регион",
    "IMEI"
}

async def validate_excel_file(profile_id: int, file: UploadFile, contents: bytes):
    if not isinstance(profile_id, int) or profile_id <= 0:
        raise BadRequestException("Parameter 'profile_id' can not be less than or equal 0")

    if not file.filename.endswith(".xlsx"):
        raise BadRequestException("File must be in .xlsx format")

    if not contents:
        raise BadRequestException("Uploaded file is empty or corrupted")

    # TODO: только чтение файла уже занимает 17 секунд поэтому не эффективно для валидации
    #  читать файл, можно ее прикрутить в парсере данных из файла, где мы и читаем их для операций.
    # try:
    #     wb = load_workbook(BytesIO(contents), read_only=True)
    # except Exception as e:
    #     raise BadRequestException(f"Failed to read Excel file: {str(e)}")
    #
    # if REQUIRED_SHEET_NAME not in wb.sheetnames:
    #     raise BadRequestException(f"Sheet '{REQUIRED_SHEET_NAME}' not found in Excel file")
    #
    # ws = wb[REQUIRED_SHEET_NAME]
    # first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)

    # if not first_row:
    #     raise BadRequestException("Sheet is empty")
    #
    # actual_columns = set(filter(None, first_row))
    # missing = REQUIRED_COLUMNS - actual_columns

    # if missing:
    #     raise BadRequestException(f"Missing required columns: {', '.join(missing)}")

